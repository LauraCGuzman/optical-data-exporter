"""
Genera los archivos Excel sintéticos para el demo del Exportador de Datos Ópticos.

Ejecutar una sola vez para crear los archivos en demo/:
    python demo/create_demo_files.py

Los archivos generados son:
    demo/refl_demo.xlsx       — hoja de medición de reflectancia sintética
    demo/trans_pv_demo.xlsx   — hoja de transmitancia PV sintética
    demo/trans_csp_demo.xlsx  — hoja de transmitancia CSP sintética
    demo/master_demo.xlsx     — tabla maestra de destino (vacía con cabeceras)
"""

import random
import math
from datetime import date, timedelta
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

DEMO_DIR = Path(__file__).parent
random.seed(42)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _thin_border():
    thin = Side(style="thin")
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def _header_fill():
    return PatternFill(start_color="2E5090", end_color="2E5090", fill_type="solid")


def _spectrum_reflectance(n=451, base=0.92, noise=0.02):
    """Espectro de reflectancia sintético: alta reflectividad con caída suave en UV."""
    values = []
    for i in range(n):
        wavelength_factor = 1 - 0.15 * math.exp(-i / 50)  # caída suave en UV
        val = base * wavelength_factor + random.gauss(0, noise)
        values.append(round(max(0.0, min(1.0, val)), 4))
    return values


def _spectrum_transmittance(n=451, base=0.91, noise=0.015):
    """Espectro de transmitancia sintético para vidrio solar."""
    values = []
    for i in range(n):
        wavelength_factor = 1 - 0.10 * math.exp(-i / 60)
        val = base * wavelength_factor + random.gauss(0, noise)
        values.append(round(max(0.0, min(1.0, val)), 4))
    return values


def _random_date(start_year=2023, end_year=2024):
    start = date(start_year, 1, 1)
    end = date(end_year, 12, 31)
    delta = (end - start).days
    return start + timedelta(days=random.randint(0, delta))


# ---------------------------------------------------------------------------
# Hoja de medición de REFLECTANCIA
# ---------------------------------------------------------------------------

def create_reflectance_sheet(filename: Path, sample_id: int = 1, duration_hours: float = 0):
    wb = Workbook()
    ws = wb.active
    ws.title = f"DEMO-REFL-{sample_id:03d}"

    # -- Metadatos de muestra (columna C) -----------------------------------
    metadata = {
        "C5":  f"DEMO-{sample_id:03d}",                          # Sample Code
        "C7":  "SolarMirrors GmbH",                               # Manufacturer
        "C8":  f"OPT-{sample_id:04d}",                            # OPAC Code
        "C9":  "Silver-coated glass reflector",                   # Material Code
        "C10": "Demo CSP Project",                                # Project
        "C11": _random_date().strftime("%d/%m/%Y"),               # Measurement Date
        "C15": "Outdoor Test Field A",                            # Site/Test
        "C16": f"Position {sample_id}",                           # Position/Observation
        "C17": 4,                                                  # Thickness (mm)
        "C18": "1000x1000",                                        # Size (XxY mm)
        "C19": _random_date(2022, 2023).strftime("%d/%m/%Y"),     # Date Exposed
        "C20": _random_date(2023, 2024).strftime("%d/%m/%Y"),     # Date Collected
        "C21": round(duration_hours, 1),                          # Duration (hours)
        "C22": round(duration_hours / 24 / 30.5, 2),             # Duration (months)
        "C24": "Synthetic demo sample — no real data",            # Comments
        "D18": 1000,                                               # Size Y
        "G14": random.randint(0, 3),                              # Corrosion spots
        "G15": round(random.uniform(0, 0.5), 3),                 # Max original edge corrosion
        "G16": round(random.uniform(0, 0.3), 3),                 # Max cut edge corrosion
        "G24": random.randint(0, 2),                              # Blisters
        "G25": 2,                                                  # Cut edges
    }

    # Parámetros ópticos hemiesféricos
    astm = round(random.uniform(0.920, 0.945), 3)
    iso  = round(astm - random.uniform(0.002, 0.010), 3)
    h660 = round(random.uniform(0.930, 0.955), 3)
    d660 = round(random.uniform(0.910, 0.935), 3)
    std  = round(random.uniform(0.001, 0.004), 3)

    optical = {
        "C28": iso,   "D28": std,   # ISO
        "C30": astm,  "D30": std,   # ASTM
        "C31": h660,  "E31": std,   # H660
        "C35": d660,  "D35": std,   # Direct660
        "C124": "±",
    }

    for cell_ref, value in {**metadata, **optical}.items():
        ws[cell_ref] = value

    # Espectro en AJ79:AJ529 (451 puntos)
    spectrum = _spectrum_reflectance(n=451, base=astm)
    for i, val in enumerate(spectrum):
        ws.cell(row=79 + i, column=36, value=val)  # columna AJ = 36

    wb.save(filename)
    print(f"  ✓ {filename.name}")


# ---------------------------------------------------------------------------
# Hoja de medición de TRANSMITANCIA PV
# ---------------------------------------------------------------------------

def create_transmittance_pv_sheet(filename: Path, sample_id: int = 1):
    wb = Workbook()
    ws = wb.active
    ws.title = f"DEMO-TRANS-PV-{sample_id:03d}"

    metadata = {
        "C5":  f"GLASS-PV-{sample_id:03d}",
        "C7":  "SolarGlass AG",
        "C8":  f"PV-{sample_id:04d}",
        "C9":  "Low-iron tempered glass",
        "C10": "Demo PV Project",
        "C11": _random_date().strftime("%d/%m/%Y"),
        "C15": "PV Test Bench B",
        "C16": f"Unit {sample_id}",
        "C17": 3.2,
        "C18": "1650x1000",
        "C19": _random_date(2022, 2023).strftime("%d/%m/%Y"),
        "C20": _random_date(2023, 2024).strftime("%d/%m/%Y"),
        "C21": round(random.uniform(0, 4380), 1),
        "C22": "",
        "C22_comments": "Synthetic demo — no real data",
    }

    trans = round(random.uniform(0.910, 0.935), 3)
    std   = round(random.uniform(0.001, 0.003), 3)

    for cell_ref, value in metadata.items():
        if not cell_ref.endswith("_comments"):
            ws[cell_ref] = value
    ws["C22"] = "Synthetic demo — no real data"
    ws["C30"] = trans
    ws["C124"] = "±"
    ws["D30"] = std

    # Espectro en AJ79:AJ529
    spectrum = _spectrum_transmittance(n=451, base=trans)
    for i, val in enumerate(spectrum):
        ws.cell(row=79 + i, column=36, value=val)

    wb.save(filename)
    print(f"  ✓ {filename.name}")


# ---------------------------------------------------------------------------
# Hoja de medición de TRANSMITANCIA CSP
# ---------------------------------------------------------------------------

def create_transmittance_csp_sheet(filename: Path, sample_id: int = 1):
    wb = Workbook()
    ws = wb.active
    ws.title = f"DEMO-TRANS-CSP-{sample_id:03d}"

    ws["C5"]  = f"TUBE-CSP-{sample_id:03d}"
    ws["C7"]  = "HeatGlass SL"
    ws["C8"]  = f"CSP-{sample_id:04d}"
    ws["C9"]  = "Borosilicate glass envelope"
    ws["C10"] = "Demo CSP Project"
    ws["C11"] = _random_date().strftime("%d/%m/%Y")
    ws["C15"] = "Parabolic Trough Test Loop"
    ws["C16"] = f"Row {sample_id}, Position 3"
    ws["C17"] = 3.0
    ws["C18"] = "4060x125"
    ws["C19"] = _random_date(2022, 2023).strftime("%d/%m/%Y")
    ws["C20"] = _random_date(2023, 2024).strftime("%d/%m/%Y")
    ws["C21"] = round(random.uniform(0, 8760), 1)
    ws["C22"] = "Synthetic demo — no real data"

    trans = round(random.uniform(0.895, 0.920), 3)
    ws["C30"] = trans
    ws["C124"] = "±"
    ws["D30"] = round(random.uniform(0.001, 0.003), 3)

    # Espectro en AM80:AM530 (451 puntos, columna AM = 39)
    spectrum = _spectrum_transmittance(n=451, base=trans, noise=0.012)
    for i, val in enumerate(spectrum):
        ws.cell(row=80 + i, column=39, value=val)

    wb.save(filename)
    print(f"  ✓ {filename.name}")


# ---------------------------------------------------------------------------
# Tabla maestra de destino (vacía con cabeceras)
# ---------------------------------------------------------------------------

def create_master_table(filename: Path):
    wb = Workbook()

    # --- Hoja ReflectorsALL ---
    ws_refl = wb.active
    ws_refl.title = "ReflectorsALL"

    refl_headers = [
        (1,  "Sample Code"),
        (2,  "OPAC Code"),
        (3,  "Material Code"),
        (4,  "Manufacturer"),
        (5,  "Project"),
        (6,  "Site/Test"),
        (8,  "Position/Observation"),
        (13, "Thickness"),
        (14, "Size X"),
        (15, "Size Y"),
        (16, "Cut Edges"),
        (17, "Measurement Date"),
        (18, "Date Exposed"),
        (19, "Date Collected"),
        (20, "Exposure duration (hours)"),   # nombre exacto que usa Addlosses
        (21, "Duration (months)"),
        (23, "Hemispherical ASTM"),           # cabeceras exactas que usa Addlosses
        (24, "±"),
        (25, "Hemispherical ASTM std"),
        (29, "Hemispherical ISO"),
        (30, "±"),
        (31, "Hemispherical ISO std"),
        (35, "Hemispherical 660"),
        (36, "±"),
        (37, "Hemispherical 660 std"),
        (41, "ρλ,φ phi"),                    # Direct / espectral
        (42, "±"),
        (43, "ρλ,φ phi std"),
        (47, "Comments"),
        (48, "Corrosion spots"),
        (50, "Edge corrosion orig."),
        (51, "Edge corrosion cut"),
        (59, "Blisters"),
        (63, "Spectrum start (250nm)"),
    ]

    header_font  = Font(bold=True, color="FFFFFF")
    header_fill  = PatternFill(start_color="2E5090", end_color="2E5090", fill_type="solid")
    header_align = Alignment(horizontal="center", wrap_text=True)

    for col, label in refl_headers:
        cell = ws_refl.cell(row=1, column=col, value=label)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = header_align

    ws_refl.row_dimensions[1].height = 30

    # Fila de ejemplo (duración 0 — valor inicial, necesario para Addlosses)
    example_row = {
        1:  "DEMO-001",
        2:  "OPT-0001",
        3:  "Silver-coated glass reflector",
        4:  "SolarMirrors GmbH",
        5:  "Demo CSP Project",
        6:  "Outdoor Test Field A",
        8:  "Position 1",
        13: 4,
        14: 1000,
        15: 1000,
        20: 0,
        21: 0,
        17: "15/01/2023",
        23: 0.935,
        24: "±",
        25: 0.002,
        29: 0.928,
        30: "±",
        31: 0.002,
        35: 0.942,
        36: "±",
        37: 0.002,
        41: 0.931,
        42: "±",
        43: 0.002,
        47: "Initial measurement (t=0)",
    }
    for col, val in example_row.items():
        ws_refl.cell(row=2, column=col, value=val)

    # --- Hoja TransmittanceALL ---
    ws_trans = wb.create_sheet("TransmittanceALL")

    trans_headers = [
        (1,  "Sample Code"),
        (2,  "OPAC Code"),
        (3,  "Material Code"),
        (4,  "Manufacturer"),
        (5,  "Project"),
        (6,  "Site/Test"),
        (7,  "Position/Observation"),
        (12, "Thickness"),
        (13, "Size"),
        (14, "Measurement Date"),
        (15, "Date Exposed"),
        (16, "Date Collected"),
        (17, "Exposure Duration"),
        (18, "Solar-weighted Transmittance"),
        (19, "±"),
        (20, "Transmittance Uncertainty"),
        (27, "Comments"),
        (38, "Spectrum start (250nm)"),
    ]

    for col, label in trans_headers:
        cell = ws_trans.cell(row=1, column=col, value=label)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = header_align

    ws_trans.row_dimensions[1].height = 30

    wb.save(filename)
    print(f"  ✓ {filename.name}")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    DEMO_DIR.mkdir(exist_ok=True)
    print("Generando archivos demo sintéticos...\n")

    # Tres hojas de reflectancia: t=0, t=1 año, t=2 años
    create_reflectance_sheet(DEMO_DIR / "refl_demo.xlsx",       sample_id=1, duration_hours=0)
    create_reflectance_sheet(DEMO_DIR / "refl_demo_1y.xlsx",    sample_id=1, duration_hours=8760)
    create_reflectance_sheet(DEMO_DIR / "refl_demo_2y.xlsx",    sample_id=1, duration_hours=17520)

    # Transmitancia PV y CSP
    create_transmittance_pv_sheet(DEMO_DIR / "trans_pv_demo.xlsx",  sample_id=1)
    create_transmittance_csp_sheet(DEMO_DIR / "trans_csp_demo.xlsx", sample_id=1)

    # Tabla maestra de destino
    create_master_table(DEMO_DIR / "master_demo.xlsx")

    print("\n✅ Demo files created in demo/")
    print("   Run: python src/main.py --config config/config_demo.json")
