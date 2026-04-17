# -*- coding: utf-8 -*-
"""
Modulo de escritura de datos a Excel
Escribe los datos al archivo Excel destino
"""

import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from copy import copy
import math

FMT_3_DEC = '0.000'
FMT_4_DEC = '0.0000'


def _copiar_estilo(celda_origen, celda_destino):
    """
    Copia todos los atributos de estilo de una celda a otra,
    sin tocar el valor.
    """
    if celda_origen.has_style:
        celda_destino.font          = copy(celda_origen.font)
        celda_destino.border        = copy(celda_origen.border)
        celda_destino.fill          = copy(celda_origen.fill)
        celda_destino.number_format = celda_origen.number_format
        celda_destino.protection    = copy(celda_origen.protection)
        celda_destino.alignment     = copy(celda_origen.alignment)


class ExcelWriter:
    """Escribe datos a archivos Excel destino"""

    def __init__(self):
        self.workbook = None
        self.filepath = None

    def abrir_excel_destino(self, path):
        """
        Abre el archivo Excel destino.
        Preserva macros solo si el archivo es .xlsm / .xlm.
        """
        self.filepath = path
        extension = path.split(".")[-1].lower()
        if extension in ("xlsm", "xlm"):
            self.workbook = load_workbook(path, keep_vba=True)
        else:
            self.workbook = load_workbook(path)
        return self.workbook

    def encontrar_ultima_fila(self, nombre_hoja):
        """
        Encuentra la siguiente fila disponible en la hoja.
        Recorre desde el final para mayor velocidad.
        """
        if self.workbook is None:
            raise ValueError("No hay workbook abierto")
        sheet = self.workbook[nombre_hoja]
        for row in range(sheet.max_row, 0, -1):
            if sheet.cell(row, 1).value is not None:
                return row + 1
        return 2

    def desactivar_filtros(self, nombre_hoja):
        """Desactiva los filtros de la hoja si existen."""
        if self.workbook is None:
            raise ValueError("No hay workbook abierto")
        sheet = self.workbook[nombre_hoja]
        if sheet.auto_filter and sheet.auto_filter.ref:
            sheet.auto_filter.ref = None

    def _aplicar_estilos_fila(self, sheet, fila_nueva, fila_referencia, max_col):
        """
        Copia el estilo de cada celda de fila_referencia a fila_nueva.
        Se llama ANTES de escribir los valores, para que la escritura posterior
        pueda sobreescribir solo el number_format donde sea necesario.
        """
        for col in range(1, max_col + 1):
            origen  = sheet.cell(row=fila_referencia, column=col)
            destino = sheet.cell(row=fila_nueva,      column=col)
            _copiar_estilo(origen, destino)

    def escribir_datos(self, nombre_hoja, fila, datos, config_celdas):
        """
        Escribe datos en una fila especifica.

        1. Copia estilos de la fila anterior (alineacion, fuente, bordes,
           number_format de fecha, etc.) para que la nueva fila sea identica
           visualmente a las existentes.
        2. Escribe los valores.
        3. Aplica formato de decimales solo a los campos numericos propios:
             - Celdas individuales float -> 0.000
             - Rangos float              -> 0.0000
        """
        if self.workbook is None:
            raise ValueError("No hay workbook abierto")
        sheet = self.workbook[nombre_hoja]

        # -- Paso 1: copiar estilos de la fila anterior ----------------------
        fila_ref = fila - 1
        if fila_ref >= 1:
            cols_usadas = []
            for cc in config_celdas:
                cd = cc.get('columna_destino')
                if isinstance(cd, list):
                    cols_usadas.extend(cd)
                elif isinstance(cd, int):
                    cols_usadas.append(cd)
            max_col = (max(cols_usadas) + 6) if cols_usadas else sheet.max_column
            self._aplicar_estilos_fila(sheet, fila, fila_ref, max_col)

        # -- Paso 2: escribir valores y formatos numericos -------------------
        for config_celda in config_celdas:
            concepto        = config_celda['concepto']
            columna_destino = config_celda.get('columna_destino')
            tipo            = config_celda.get('tipo', 'celda')

            if concepto not in datos or columna_destino is None:
                continue

            valor = datos[concepto]

            if isinstance(columna_destino, list):
                for col in columna_destino:
                    celda = sheet.cell(row=fila, column=col, value=valor)
                    if isinstance(valor, float):
                        celda.number_format = FMT_3_DEC

            elif tipo == 'rango' and isinstance(valor, list):
                for i, v in enumerate(valor):
                    celda = sheet.cell(row=fila, column=columna_destino + i, value=v)
                    if isinstance(v, float):
                        celda.number_format = FMT_4_DEC

            else:
                celda = sheet.cell(row=fila, column=columna_destino, value=valor)
                if isinstance(valor, float):
                    celda.number_format = FMT_3_DEC

    # -------------------------------------------------------------------------
    # Reimplementacion del macro VBA "Addlosses"
    # -------------------------------------------------------------------------

    def ejecutar_addlosses(self, nombre_hoja, filas_nuevas):
        """
        Reimplementacion en Python del macro VBA 'Addlosses'.

        Para cada fila nueva (duracion > 0) busca el valor inicial
        (duracion = 0) del mismo sample name y escribe:
          col_metrica + 3  -> diferencia (valor_actual - valor_inicial)
          col_metrica + 4  -> simbolo '+-'
          col_metrica + 5  -> incertidumbre combinada sqrt(unc^2 + unc_ini^2)

        Metricas localizadas por cabecera:
          Hemispherical ASTM, Hemispherical ISO, Hemispherical 660, ρλ,φ phi
        """
        if self.workbook is None:
            raise ValueError("No hay workbook abierto")
        sheet = self.workbook[nombre_hoja]

        cabeceras_buscadas = {
            'dur':  'Exposure duration (hours)',
            'spec': 'ρλ,φ',
            'astm': 'Hemispherical ASTM',
            'iso':  'Hemispherical ISO',
            'hem':  'Hemispherical 660',
        }
        col_map = {}
        for col in range(1, sheet.max_column + 1):
            cabecera = sheet.cell(1, col).value
            if cabecera is None:
                continue
            for key, texto in cabeceras_buscadas.items():
                if key not in col_map and texto.lower() in str(cabecera).lower():
                    col_map[key] = col

        if 'dur' not in col_map:
            print("  Addlosses: columna de duracion no encontrada. Saltando.")
            return

        dur_col  = col_map['dur']
        metricas = {k: col_map[k] for k in ('spec', 'astm', 'iso', 'hem') if k in col_map}

        if not metricas:
            print("  Addlosses: columnas de metricas no encontradas. Saltando.")
            return

        sample_col      = 1
        filas_iniciales = {}
        for row in range(2, sheet.max_row + 1):
            dur_val = sheet.cell(row, dur_col).value
            sample  = sheet.cell(row, sample_col).value
            if dur_val == 0 and sample:
                filas_iniciales[sample] = row

        filas_procesadas = 0
        for fila in filas_nuevas:
            if sheet.cell(fila, dur_col).value == 0:
                continue

            sample = sheet.cell(fila, sample_col).value
            if not sample or sample not in filas_iniciales:
                print(f"  Addlosses: sin valor inicial para '{sample}' (fila {fila}).")
                continue

            ini_row = filas_iniciales[sample]

            for met_col in metricas.values():
                val_actual = sheet.cell(fila,    met_col).value
                val_ini    = sheet.cell(ini_row, met_col).value
                if val_actual is None or val_ini is None:
                    continue

                drop = val_actual - val_ini
                cd = sheet.cell(row=fila, column=met_col + 3, value=drop)
                cd.number_format = FMT_3_DEC

                unc_actual = sheet.cell(fila,    met_col + 2).value
                unc_ini    = sheet.cell(ini_row, met_col + 2).value
                if unc_actual is not None and unc_ini is not None:
                    unc_comb = math.sqrt(unc_actual ** 2 + unc_ini ** 2)
                    sheet.cell(row=fila, column=met_col + 4, value='+-')
                    cu = sheet.cell(row=fila, column=met_col + 5, value=unc_comb)
                    cu.number_format = FMT_3_DEC

            filas_procesadas += 1

        print(f"  Addlosses: {filas_procesadas} fila(s) procesada(s).")

    # -------------------------------------------------------------------------

    def guardar_y_cerrar(self):
        """Guarda y cierra el workbook"""
        if self.workbook:
            self.workbook.save(self.filepath)
            self.workbook.close()
            self.workbook = None

    def cerrar_sin_guardar(self):
        """Cierra el workbook sin guardar cambios"""
        if self.workbook:
            self.workbook.close()
            self.workbook = None