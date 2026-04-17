# -*- coding: utf-8 -*-
"""
Módulo de lectura de datos desde Excel
Lee los datos del archivo Excel fuente según la configuración
"""

import openpyxl
from openpyxl import load_workbook
import re


class ExcelReader:
    """Lee datos de archivos Excel fuente"""
    
    def __init__(self):
        self.workbook = None
        self.filepath = None
        self._cache_hojas = {}  # Cache de valores de hoja: {nombre_hoja: {celda: valor}}
    
    def abrir_workbook(self, path):
        """
        Abre un workbook de Excel
        
        Args:
            path: Path del archivo Excel
        
        Returns:
            workbook object
        """
        self.filepath = path
        # read_only=True acelera enormemente la lectura cuando no se va a escribir
        self.workbook = load_workbook(path, data_only=True, read_only=True)
        self._cache_hojas = {}
        return self.workbook
    
    def get_sheet_names(self):
        """Obtiene lista de nombres de hojas del workbook actual"""
        if self.workbook is None:
            raise ValueError("No hay workbook abierto")
        return self.workbook.sheetnames
    
    def _precargar_hoja(self, sheet):
        """
        Precarga todos los valores de una hoja en un dict {coord: valor}.
        Evita acceder celda a celda al sheet (lento en modo normal/read_only).
        """
        nombre = sheet.title
        if nombre in self._cache_hojas:
            return self._cache_hojas[nombre]
        
        cache = {}
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value is not None:
                    cache[cell.coordinate] = cell.value
        self._cache_hojas[nombre] = cache
        return cache

    def leer_datos_hoja(self, nombre_hoja, config_celdas):
        """
        Lee datos de una hoja según la configuración
        
        Args:
            nombre_hoja: Nombre de la hoja a leer
            config_celdas: Lista de configuración de celdas (desde config.json)
        
        Returns:
            dict con los datos leídos {concepto: valor}
        """
        if self.workbook is None:
            raise ValueError("No hay workbook abierto")
        
        sheet = self.workbook[nombre_hoja]
        cache = self._precargar_hoja(sheet)
        datos = {}

        # Precalcular mapa concepto→config para evitar bucles anidados
        concepto_map = {c['concepto']: c for c in config_celdas}
        
        for config_celda in config_celdas:
            origen = config_celda['origen']
            concepto = config_celda['concepto']
            tipo = config_celda.get('tipo', 'celda')
            
            if tipo == 'rango':
                valor = self._leer_rango_cache(sheet, origen)
            else:
                valor = cache.get(origen)
                
                # Formatear fechas a dd/mm/yyyy
                if concepto in ("Measurement Date", "Date Exposed", "Date Collected"):
                    from datetime import datetime
                    if isinstance(valor, datetime):
                        valor = valor.strftime('%d/%m/%Y')
                    elif isinstance(valor, str) and ' ' in valor:
                        try:
                            parts = valor.split(' ')[0].split('-')
                            if len(parts) == 3:
                                valor = f"{parts[2]}/{parts[1]}/{parts[0]}"
                        except Exception:
                            pass
                
                # Transformaciones de dimensiones
                if concepto == "Size X" and isinstance(valor, str) and 'x' in str(valor).lower():
                    valor = self._dividir_dimensiones(valor)[0]
                elif concepto == "Size Y":
                    size_x_config = concepto_map.get('Size X')
                    if size_x_config:
                        size_x_valor = cache.get(size_x_config['origen'])
                        if isinstance(size_x_valor, str) and 'x' in str(size_x_valor).lower():
                            valor = self._dividir_dimensiones(size_x_valor)[1]
                
                # Conversión de duración
                if concepto == "Duration (months)" and not valor:
                    horas_config = concepto_map.get('Duration (hours)')
                    if horas_config:
                        horas = cache.get(horas_config['origen'])
                        if horas and isinstance(horas, (int, float)):
                            valor = horas / 24 / 30.5
                
                elif concepto == "Duration (hours)" and not valor:
                    meses_config = concepto_map.get('Duration (months)')
                    if meses_config:
                        meses = cache.get(meses_config['origen'])
                        if meses and isinstance(meses, (int, float)):
                            valor = meses * 24 * 30.5

            # Redondear números: 4 decimales para rangos, 3 para el resto
            if tipo == 'rango' and isinstance(valor, list):
                valor = [round(float(v), 4) if isinstance(v, (int, float)) else v for v in valor]
            elif isinstance(valor, (int, float)):
                valor = round(float(valor), 3)
            
            datos[concepto] = valor
        
        return datos
    
    def _leer_rango_cache(self, sheet, rango_ref):
        """Lee un rango iterando filas (compatible con read_only)"""
        try:
            if ':' not in rango_ref:
                # Celda individual tratada como rango
                val = None
                for row in sheet.iter_rows(min_row=1):
                    for cell in row:
                        if cell.coordinate == rango_ref:
                            val = cell.value
                            break
                return [val] if val is not None else []
            
            valores = []
            for fila in sheet[rango_ref]:
                for celda in fila:
                    if celda.value is not None:
                        valores.append(celda.value)
            return valores
        except Exception as e:
            raise ValueError(f"Error al leer rango {rango_ref}: {e}")

    def _leer_celda(self, sheet, celda_ref):
        """Lee el valor de una celda individual (mantener por compatibilidad)"""
        try:
            return sheet[celda_ref].value
        except Exception as e:
            raise ValueError(f"Error al leer celda {celda_ref}: {e}")
    
    def _leer_rango(self, sheet, rango_ref):
        """Lee el valor de un rango de celdas y retorna una lista"""
        return self._leer_rango_cache(sheet, rango_ref)
    
    def _dividir_dimensiones(self, dimension_str):
        """
        Divide una string de dimensiones "XxY" en dos valores
        """
        if not isinstance(dimension_str, str):
            return dimension_str, None
        
        match = re.search(r'(\d+\.?\d*)\s*[xX]\s*(\d+\.?\d*)', str(dimension_str))
        if match:
            x = float(match.group(1))
            y = float(match.group(2))
            if x.is_integer():
                x = int(x)
            if y.is_integer():
                y = int(y)
            return x, y
        
        return dimension_str, None
    
    def cerrar(self):
        """Cierra el workbook"""
        if self.workbook:
            self.workbook.close()
            self.workbook = None
            self._cache_hojas = {}
